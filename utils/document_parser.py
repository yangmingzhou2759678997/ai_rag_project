import os
from io import BytesIO

from docx import Document as WordDocument
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text_file(file_content: bytes) -> str:
    """读取UTF-8编码的TXT或Markdown文件。"""
    try:
        return file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValueError("TXT或Markdown文件必须使用UTF-8编码")


def extract_pdf(file_content: bytes) -> str:
    """提取带文字层PDF中的文字，并保留简单的页面标记。"""
    pdf_reader = PdfReader(BytesIO(file_content))
    page_text_list = []

    for page_number, page in enumerate(pdf_reader.pages, start=1):
        page_text = page.extract_text(extraction_mode="layout")

        if page_text and page_text.strip():
            page_text_list.append(f"第{page_number}页：\n{page_text.strip()}")

    return "\n\n".join(page_text_list)


def get_docx_table_text(table: Table) -> str:
    """把DOCX表格按行转换成简单文本。"""
    row_text_list = []

    for row in table.rows:
        cell_text_list = []

        for cell in row.cells:
            cell_text_list.append(cell.text.strip())

        if any(cell_text_list):
            row_text_list.append(" | ".join(cell_text_list))

    if not row_text_list:
        return ""

    return "【表格】\n" + "\n".join(row_text_list)


def extract_docx(file_content: bytes) -> str:
    """使用python-docx高层API，按原顺序提取段落和表格。"""
    word_document = WordDocument(BytesIO(file_content))
    content_list = []

    for block in word_document.iter_inner_content():
        text = ""

        if isinstance(block, Paragraph):
            text = block.text.strip()
        elif isinstance(block, Table):
            text = get_docx_table_text(block)

        if text:
            content_list.append(text)

    return "\n\n".join(content_list)


def extract_xlsx(file_content: bytes) -> str:
    """按工作表和行读取XLSX单元格中的值。"""
    workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    sheet_text_list = []

    try:
        for sheet in workbook.worksheets:
            row_text_list = []

            for row in sheet.iter_rows(values_only=True):
                cell_text_list = []

                for value in row:
                    if value is None:
                        cell_text_list.append("")
                    else:
                        cell_text_list.append(str(value).strip())

                while cell_text_list and not cell_text_list[-1]:
                    cell_text_list.pop()

                if any(cell_text_list):
                    row_text_list.append(" | ".join(cell_text_list))

            if row_text_list:
                sheet_text = f"【工作表：{sheet.title}】\n" + "\n".join(row_text_list)
                sheet_text_list.append(sheet_text)
    finally:
        workbook.close()

    return "\n\n".join(sheet_text_list)


def extract_document(file_name: str, file_content: bytes) -> str:
    """根据文件扩展名调用对应的文本提取函数。"""
    file_extension = os.path.splitext(file_name)[1].lower()

    if file_extension == ".txt" or file_extension == ".md":
        return extract_text_file(file_content)

    if file_extension == ".pdf":
        return extract_pdf(file_content)

    if file_extension == ".docx":
        return extract_docx(file_content)

    if file_extension == ".xlsx":
        return extract_xlsx(file_content)

    raise ValueError("当前只支持txt、md、pdf、docx、xlsx文件")
