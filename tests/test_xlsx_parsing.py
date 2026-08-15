import unittest
from io import BytesIO

from openpyxl import Workbook

from utils.document_parser import extract_document
from utils.text_splitter import split_text


def workbook_to_bytes(workbook):
    file_buffer = BytesIO()
    workbook.save(file_buffer)
    return file_buffer.getvalue()


class TestXlsxParsing(unittest.TestCase):
    def test_multiple_sheets_and_rows_are_extracted(self):
        workbook = Workbook()
        employee_sheet = workbook.active
        employee_sheet.title = "员工"
        employee_sheet.append(["姓名", "部门", "年假"])
        employee_sheet.append(["张三", "研发", 5])
        rule_sheet = workbook.create_sheet("制度")
        rule_sheet.append(["规则", "每年复核"])
        text = extract_document("employees.xlsx", workbook_to_bytes(workbook))
        self.assertIn("【工作表：员工】", text)
        self.assertIn("姓名 | 部门 | 年假", text)
        self.assertIn("张三 | 研发 | 5", text)
        self.assertIn("【工作表：制度】", text)
        self.assertLess(text.index("【工作表：员工】"), text.index("【工作表：制度】"))

    def test_empty_middle_cell_keeps_column_position(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["A", None, "C"])
        text = extract_document("empty-cell.xlsx", workbook_to_bytes(workbook))
        self.assertIn("A |  | C", text)

    def test_empty_sheets_do_not_create_text(self):
        workbook = Workbook()
        workbook.create_sheet("另一个空表")
        text = extract_document("empty.xlsx", workbook_to_bytes(workbook))
        self.assertEqual(text, "")

    def test_xlsx_text_uses_common_splitter(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["项目", "说明"])
        sheet.append(["知识库", "逐行读取单元格内容"])
        text = extract_document("knowledge.xlsx", workbook_to_bytes(workbook))
        chunks = split_text(text, 350, 50)
        self.assertTrue(chunks)
        self.assertIn("逐行读取单元格内容", "".join(chunks))


if __name__ == "__main__":
    unittest.main()
